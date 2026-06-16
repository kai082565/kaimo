import os
import sys
import threading
import webbrowser
from datetime import date
from functools import wraps

from flask import (Flask, render_template, request, redirect,
                   url_for, jsonify, flash, send_file, session)
from werkzeug.security import check_password_hash

import database as db


def _resource(rel):
    # PyInstaller 打包後資源在 _MEIPASS；開發時在原始目錄
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)


app = Flask(__name__,
            template_folder=_resource('templates'),
            static_folder=_resource('static'))
app.secret_key = "pawnshop-2024"
app.config['TEMPLATES_AUTO_RELOAD'] = True


# ── 認證工具 ──────────────────────────────────────────

# 業務角色允許存取的頁面
_STAFF_ALLOWED = {
    'tickets_monthly', 'tickets_unpaid',
    'new_ticket', 'ticket_detail',
    'customers', 'customer_detail', 'edit_customer',
    'api_interest', 'logout', 'static', 'settings',
    'blacklist',
}


def _current_staff_id():
    """若目前登入者是業務，回傳其 staff_id；否則回傳 None（不篩選）"""
    if session.get('role') == '業務':
        return session.get('staff_id')
    return None


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


@app.before_request
def check_auth():
    if request.endpoint == 'static':
        return
    if not db.has_any_user():
        if request.endpoint != 'setup':
            return redirect(url_for('setup'))
        return
    if request.endpoint in ('login', 'logout', 'setup'):
        return
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('role') == '業務' and request.endpoint not in _STAFF_ALLOWED:
        return redirect(url_for('tickets_monthly'))


# ── 認證路由 ──────────────────────────────────────────

@app.route('/setup', methods=['GET', 'POST'])
def setup():
    if db.has_any_user():
        return redirect(url_for('login'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        if not username or not password:
            flash('帳號和密碼不能為空', 'danger')
            return render_template('auth/setup.html')
        if password != confirm_password:
            flash('兩次密碼不一致，請重新輸入', 'danger')
            return render_template('auth/setup.html')
        if len(password) < 4:
            flash('密碼至少需要 4 個字元', 'danger')
            return render_template('auth/setup.html')
        db.create_user({'username': username, 'password': password, 'role': '老闆'})
        flash('老闆帳號建立成功，請登入！', 'success')
        return redirect(url_for('login'))
    return render_template('auth/setup.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        role = session.get('role')
        return redirect(url_for('tickets_monthly') if role == '業務' else url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = db.get_user_by_username(username)
        if user and check_password_hash(user['password_hash'], password):
            session['user_id']  = user['id']
            session['username'] = user['username']
            session['role']     = user['role']
            session['staff_id'] = user['staff_id']
            if user['role'] == '業務':
                return redirect(url_for('tickets_monthly'))
            return redirect(url_for('dashboard'))
        flash('帳號或密碼錯誤', 'danger')
    return render_template('auth/login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── 工具 ──────────────────────────────────────────────

def _today():
    return date.today().isoformat()


# ── 儀表板 ────────────────────────────────────────────

@app.route("/")
def dashboard():
    stats = db.get_dashboard_stats()
    due_soon = db.get_due_soon_tickets(14)
    recent = db.get_recent_history(8)
    settings = db.get_settings()
    return render_template("dashboard.html",
                           stats=stats,
                           due_soon=due_soon,
                           recent=recent,
                           settings=settings,
                           today=_today())


# ── 當票 ──────────────────────────────────────────────

@app.route("/tickets")
def tickets():
    status = request.args.get("status", "")
    search = request.args.get("q", "")
    rows = db.list_tickets(status or None, search or None)
    categories = db.get_categories()
    return render_template("tickets/list.html",
                           tickets=rows,
                           page_title="當票管理",
                           status_filter=status,
                           search=search,
                           categories=categories,
                           today=_today())


@app.route("/tickets/monthly")
def tickets_monthly():
    rows = db.list_tickets_monthly(staff_id=_current_staff_id())
    return render_template("tickets/list.html",
                           tickets=rows,
                           page_title="當月應收",
                           status_filter="", search="", today=_today())


@app.route("/tickets/unpaid")
def tickets_unpaid():
    rows = db.list_tickets_unpaid(staff_id=_current_staff_id())
    return render_template("tickets/list.html",
                           tickets=rows,
                           page_title="應收未收",
                           status_filter="", search="", today=_today())


@app.route("/tickets/new", methods=["GET", "POST"])
def new_ticket():
    categories = db.get_categories()
    settings = db.get_settings()
    staff_list = db.get_staff_list()

    if request.method == "POST":
        data = request.form.to_dict()
        if session.get('role') == '業務':
            data['staff_id'] = session.get('staff_id')
        cid = db.create_customer({
            "name":    data.get("customer_name", "").strip(),
            "id_card": data.get("customer_id_card", "").strip(),
            "phone":   data.get("customer_phone", "").strip(),
            "address": data.get("customer_address", "").strip(),
        })
        data["customer_id"] = cid

        tid = db.create_ticket(data)
        flash("當票開立成功！", "success")
        return redirect(url_for("ticket_detail", ticket_id=tid))

    return render_template("tickets/new.html",
                           categories=categories,
                           settings=settings,
                           staff_list=staff_list,
                           today=_today())


@app.route("/tickets/<int:ticket_id>")
def ticket_detail(ticket_id):
    ticket = db.get_ticket(ticket_id)
    if not ticket:
        flash("找不到該當票", "danger")
        return redirect(url_for("tickets_monthly"))
    if session.get('role') == '業務' and ticket.get('staff_id') != session.get('staff_id'):
        flash("無權限查看此當票", "danger")
        return redirect(url_for("tickets_monthly"))
    return render_template("tickets/detail.html",
                           ticket=ticket,
                           today=_today())


@app.route("/tickets/<int:ticket_id>/edit", methods=["GET", "POST"])
def edit_ticket(ticket_id):
    ticket = db.get_ticket(ticket_id)
    if not ticket:
        flash("找不到該當票", "danger")
        return redirect(url_for("tickets"))
    if request.method == "POST":
        db.update_ticket(ticket_id, request.form.to_dict())
        flash("當票資料已更新！", "success")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))
    categories = db.get_categories()
    staff_list = db.get_staff_list()
    customer = db.get_customer(ticket["customer_id"]) if ticket.get("customer_id") else None
    return render_template("tickets/edit.html",
                           ticket=ticket,
                           customer=customer,
                           categories=categories,
                           staff_list=staff_list,
                           today=_today())


@app.route("/tickets/<int:ticket_id>/redeem", methods=["POST"])
def redeem(ticket_id):
    calc_date = request.form.get("calc_date") or _today()
    notes = request.form.get("notes", "")
    db.redeem_ticket(ticket_id, calc_date, notes)
    flash("贖回完成！", "success")
    return redirect(url_for("ticket_detail", ticket_id=ticket_id))


@app.route("/tickets/<int:ticket_id>/renew", methods=["POST"])
def renew(ticket_id):
    new_months = int(request.form.get("new_months", 3))
    calc_date = request.form.get("calc_date") or _today()
    notes = request.form.get("notes", "")
    db.renew_ticket(ticket_id, new_months, calc_date, notes)
    flash("續當成功！", "success")
    return redirect(url_for("ticket_detail", ticket_id=ticket_id))


@app.route("/tickets/<int:ticket_id>/forfeit", methods=["POST"])
def forfeit(ticket_id):
    notes = request.form.get("notes", "")
    db.forfeit_ticket(ticket_id, notes)
    flash("已標記為流當。", "warning")
    return redirect(url_for("ticket_detail", ticket_id=ticket_id))


@app.route("/tickets/<int:ticket_id>/delete", methods=["POST"])
def delete_ticket(ticket_id):
    db.delete_ticket(ticket_id)
    return redirect(url_for("tickets"))


@app.route("/tickets/<int:ticket_id>/pay_period/<int:schedule_id>", methods=["POST"])
def pay_period(ticket_id, schedule_id):
    data = request.get_json() or {}
    db.record_period_payment(
        schedule_id,
        float(data.get("paid_amount", 0)),
        float(data.get("paid_principal", 0)),
        float(data.get("late_fee", 0)),
        data.get("notes", ""),
    )
    return jsonify({"ok": True})


@app.route("/tickets/<int:ticket_id>/repay", methods=["POST"])
def repay_principal(ticket_id):
    data = request.get_json() or {}
    schedule_id = int(data["schedule_id"]) if data.get("schedule_id") else None
    db.repay_principal(ticket_id, float(data.get("amount", 0)), data.get("notes", ""), schedule_id)
    return jsonify({"ok": True})


@app.route("/tickets/<int:ticket_id>/settle", methods=["POST"])
def settle_schedule(ticket_id):
    db.settle_ticket_schedule(ticket_id)
    return jsonify({"ok": True})


@app.route("/tickets/<int:ticket_id>/cancel_payment/<int:schedule_id>", methods=["POST"])
def cancel_payment(ticket_id, schedule_id):
    db.cancel_period_payment(schedule_id)
    return jsonify({"ok": True})


@app.route("/tickets/<int:ticket_id>/cancel_repayment/<int:schedule_id>", methods=["POST"])
def cancel_repayment(ticket_id, schedule_id):
    db.cancel_period_repayment(schedule_id)
    return jsonify({"ok": True})


# ── 客戶 ──────────────────────────────────────────────

@app.route("/customers")
def customers():
    search = request.args.get("q", "")
    rows = db.list_customers(search or None, staff_id=_current_staff_id())
    return render_template("customers/list.html",
                           customers=rows, search=search)


@app.route("/customers/<int:customer_id>")
def customer_detail(customer_id):
    customer = db.get_customer(customer_id)
    if not customer:
        flash("找不到該客戶", "danger")
        return redirect(url_for("customers"))
    if session.get('role') == '業務':
        sid = session.get('staff_id')
        if not any(t.get('staff_id') == sid for t in customer.get('tickets', [])):
            flash("無權限查看此客戶", "danger")
            return redirect(url_for("customers"))
    return render_template("customers/detail.html",
                           customer=customer, today=_today())


@app.route("/customers/<int:customer_id>/delete", methods=["POST"])
def delete_customer(customer_id):
    if session.get('role') == '業務':
        return redirect(url_for("customers"))
    db.delete_customer(customer_id)
    flash("客戶已刪除。", "success")
    return redirect(url_for("customers"))


@app.route("/customers/<int:customer_id>/edit", methods=["GET", "POST"])
def edit_customer(customer_id):
    customer = db.get_customer(customer_id)
    if not customer:
        return redirect(url_for("customers"))
    if session.get('role') == '業務':
        sid = session.get('staff_id')
        if not any(t.get('staff_id') == sid for t in customer.get('tickets', [])):
            flash("無權限編輯此客戶", "danger")
            return redirect(url_for("customers"))
    if request.method == "POST":
        db.update_customer(customer_id, request.form.to_dict())
        flash("客戶資料已更新！", "success")
        return redirect(url_for("customer_detail", customer_id=customer_id))
    return render_template("customers/form.html",
                           customer=customer, action="edit")



@app.route("/api/interest/<int:ticket_id>")
def api_interest(ticket_id):
    calc_date = request.args.get("date", _today())
    new_months = int(request.args.get("new_months", 3))
    t = db.get_ticket(ticket_id)
    if not t:
        return jsonify({"error": "not found"}), 404
    months = db.calc_months(t["pawn_date"], calc_date)
    interest = db.calc_interest(t["principal"], t["monthly_rate"], months)
    from dateutil.relativedelta import relativedelta
    new_due = (date.fromisoformat(calc_date) + relativedelta(months=new_months)).isoformat()
    return jsonify({
        "months": months,
        "interest": interest,
        "total": t["principal"] + interest,
        "principal": t["principal"],
        "new_due": new_due,
    })




# ── 業績排名 ──────────────────────────────────────────

@app.route("/performance")
def performance():
    import calendar
    year  = int(request.args.get("year",  date.today().year))
    month = int(request.args.get("month", date.today().month))
    data  = db.get_performance_ranking(year, month)
    last_day = calendar.monthrange(year, month)[1]
    return render_template("performance.html",
                           data=data, year=year, month=month,
                           last_day=last_day,
                           current_year=date.today().year,
                           current_month=date.today().month)


@app.route("/performance/export")
def performance_export():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    year  = int(request.args.get("year",  date.today().year))
    month = int(request.args.get("month", date.today().month))
    data  = db.get_performance_ranking(year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年{month:02d}月業績"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"{year} 年 {month:02d} 月  業績排名報表"
    ws["A1"].font = Font(size=13, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["名次", "姓名", "電話", "票面金額", "應收利息", "實收利息", "收款率(%)", "客戶數"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1d4ed8")
        c.alignment = Alignment(horizontal="center")

    trend_map = {"up": "↑ 上升", "down": "↓ 下降", "stable": "→ 持平"}
    for i, r in enumerate(data["ranking"], 1):
        row = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r["name"])
        ws.cell(row=row, column=3, value=r["phone"])
        ws.cell(row=row, column=4, value=int(r["principal"]))
        ws.cell(row=row, column=5, value=int(r["expected"]))
        ws.cell(row=row, column=6, value=int(r["collected"]))
        ws.cell(row=row, column=7, value=r["collection_rate"])
        ws.cell(row=row, column=8, value=r["customer_count"])

    for col, w in zip("ABCDEFGH", [6, 10, 13, 13, 12, 12, 11, 8]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"業績排名_{year}年{month:02d}月.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── 帳務管理 ──────────────────────────────────────────

@app.route("/accounts")
def accounts():
    return redirect(url_for("accounts_overdue"))


@app.route("/accounts/overdue")
def accounts_overdue():
    search   = request.args.get("q", "")
    ref_date = request.args.get("date", _today())
    data     = db.get_overdue_overview(search or None, ref_date)
    return render_template("accounts/overdue.html",
                           data=data, search=search,
                           ref_date=ref_date, today=_today())


@app.route("/api/accounts/contact/<int:schedule_id>", methods=["POST"])
def update_contact(schedule_id):
    body   = request.get_json() or {}
    status = body.get("status", "已聯絡")
    db.update_contact_status(schedule_id, status)
    return jsonify({"ok": True})


@app.route("/accounts/bad-debts")
def accounts_bad_debts():
    search = request.args.get("q", "")
    data   = db.get_bad_debts_overview(search or None)
    return render_template("accounts/bad_debts.html",
                           data=data, search=search, today=_today())


@app.route("/api/accounts/bad-debt/<int:bad_debt_id>/update", methods=["POST"])
def api_update_bad_debt(bad_debt_id):
    body = request.get_json() or {}
    db.update_bad_debt(
        bad_debt_id,
        float(body.get("recovered_amount", 0)),
        body.get("status", "未回收"),
        body.get("notes", ""),
    )
    return jsonify({"ok": True})


@app.route("/api/accounts/bad-debt/<int:bad_debt_id>/cancel", methods=["POST"])
def api_cancel_bad_debt(bad_debt_id):
    db.cancel_bad_debt(bad_debt_id)
    return jsonify({"ok": True})


# ── 總帳報表 ──────────────────────────────────────────

@app.route("/ledger")
def ledger():
    return redirect(url_for("ledger_overview"))


@app.route("/ledger/overview")
def ledger_overview():
    data = db.get_ledger_overview()
    return render_template("ledger/overview.html", data=data)


@app.route("/ledger/providers")
def ledger_providers():
    rows = db.list_fund_providers()
    staff_list = db.get_staff_list()
    return render_template("ledger/providers.html",
                           providers=rows, staff_list=staff_list)


@app.route("/ledger/providers/new", methods=["POST"])
def ledger_provider_new():
    db.create_fund_provider(request.form.to_dict())
    flash("資金提供方已新增！", "success")
    return redirect(url_for("ledger_providers"))


@app.route("/ledger/providers/<int:fp_id>/edit", methods=["POST"])
def ledger_provider_edit(fp_id):
    db.update_fund_provider(fp_id, request.form.to_dict())
    flash("資料已更新！", "success")
    return redirect(url_for("ledger_providers"))


@app.route("/ledger/providers/<int:fp_id>/delete", methods=["POST"])
def ledger_provider_delete(fp_id):
    db.delete_fund_provider(fp_id)
    flash("已刪除。", "success")
    return redirect(url_for("ledger_providers"))


@app.route("/ledger/transactions")
def ledger_transactions():
    df        = request.args.get("date_from", date.today().replace(day=1).isoformat())
    dt        = request.args.get("date_to",   _today())
    type_f    = request.args.get("type",      "全部")
    sort_desc = request.args.get("sort", "desc") != "asc"
    rows      = db.list_transactions(df, dt, type_f, sort_desc)
    staff_list = db.get_staff_list()
    return render_template("ledger/transactions.html",
                           rows=rows, staff_list=staff_list,
                           date_from=df, date_to=dt,
                           type_filter=type_f,
                           sort_desc=sort_desc,
                           today=_today())


@app.route("/ledger/transactions/new", methods=["POST"])
def ledger_transaction_new():
    db.create_transaction(request.form.to_dict())
    flash("交易記錄已新增！", "success")
    args = request.args.to_dict()
    return redirect(url_for("ledger_transactions", **args))


@app.route("/ledger/transactions/<int:tx_id>/edit", methods=["POST"])
def ledger_transaction_edit(tx_id):
    db.update_transaction(tx_id, request.form.to_dict())
    flash("交易記錄已更新！", "success")
    args = {k: v for k, v in request.args.items()}
    return redirect(url_for("ledger_transactions", **args))


@app.route("/ledger/transactions/<int:tx_id>/delete", methods=["POST"])
def ledger_transaction_delete(tx_id):
    db.delete_transaction(tx_id)
    flash("已刪除。", "success")
    args = {k: v for k, v in request.args.items()}
    return redirect(url_for("ledger_transactions", **args))


@app.route("/ledger/transactions/export")
def ledger_transactions_export():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    df     = request.args.get("date_from", date.today().replace(day=1).isoformat())
    dt     = request.args.get("date_to",   _today())
    type_f = request.args.get("type", "全部")
    rows   = db.list_transactions(df, dt, type_f, sort_desc=False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交易明細"

    headers = ["日期","項目","昨日餘額","出款金額","營運支出","收款金額","其他收入","資本額異動","本日餘額","建立者","類型","備註"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1d4ed8")
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1,  value=r['date'])
        ws.cell(row=i, column=2,  value=r['item'])
        ws.cell(row=i, column=3,  value=r['prev_balance'])
        ws.cell(row=i, column=4,  value=r.get('outgoing', 0))
        ws.cell(row=i, column=5,  value=r.get('op_expense', 0))
        ws.cell(row=i, column=6,  value=r.get('income', 0))
        ws.cell(row=i, column=7,  value=r.get('other_income', 0))
        ws.cell(row=i, column=8,  value=r.get('capital_change', 0))
        ws.cell(row=i, column=9,  value=r['curr_balance'])
        ws.cell(row=i, column=10, value=r.get('staff_name', ''))
        ws.cell(row=i, column=11, value=r['type'])
        ws.cell(row=i, column=12, value=r.get('notes', ''))

    for col, w in zip("ABCDEFGHIJKL", [12,20,13,11,11,11,11,11,13,10,10,18]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"交易明細_{df}_{dt}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/ledger/providers/export")
def ledger_providers_export():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = db.list_fund_providers()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "資金提供方"

    headers = ["時間", "名字", "收入", "支出", "固定金額", "建立者", "備註"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1d4ed8")
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r["created_at"])
        ws.cell(row=i, column=2, value=r["name"])
        ws.cell(row=i, column=3, value=r["income"])
        ws.cell(row=i, column=4, value=r["expense"])
        ws.cell(row=i, column=5, value=r["fixed_amount"])
        ws.cell(row=i, column=6, value=r.get("staff_name") or "")
        ws.cell(row=i, column=7, value=r.get("notes") or "")

    for col, w in zip("ABCDEFG", [20, 14, 12, 12, 12, 12, 20]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="資金提供方報表.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── 設定 ──────────────────────────────────────────────

@app.route("/settings", methods=["GET", "POST"])
def settings():
    if request.method == "POST":
        action = request.form.get("_action", "")
        if action == "change_password":
            old_pw = request.form.get("old_password", "")
            new_pw = request.form.get("new_password", "")
            confirm_pw = request.form.get("confirm_password", "")
            user = db.get_user(session['user_id'])
            from werkzeug.security import check_password_hash
            if not check_password_hash(user['password_hash'], old_pw):
                flash("舊密碼錯誤", "danger")
            elif new_pw != confirm_pw:
                flash("新密碼與確認密碼不符", "danger")
            elif len(new_pw) < 4:
                flash("密碼至少需要 4 個字元", "danger")
            else:
                db.update_user_password(session['user_id'], new_pw)
                flash("密碼已更新！", "success")
            return redirect(url_for("settings"))
        # 一般設定儲存
        db.save_settings(request.form.to_dict())
        categories = db.get_categories()
        with db.get_conn() as conn:
            for cat in categories:
                rate_key = f"cat_rate_{cat['id']}"
                name_key = f"cat_name_{cat['id']}"
                if rate_key in request.form:
                    conn.execute(
                        "UPDATE categories SET default_rate=?, name=? WHERE id=?",
                        (request.form[rate_key], request.form.get(name_key, cat["name"]), cat["id"])
                    )
        flash("設定已儲存！", "success")
        return redirect(url_for("settings"))
    cfg = db.get_settings()
    categories = db.get_categories()
    users = db.list_users() if session.get('role') == '老闆' else []
    staff_list = db.get_staff_list() if session.get('role') == '老闆' else []
    return render_template("settings.html", cfg=cfg, categories=categories,
                           users=users, staff_list=staff_list)


@app.route("/settings/users/new", methods=["POST"])
def settings_user_new():
    if session.get('role') != '老闆':
        return redirect(url_for('settings'))
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    role = request.form.get('role', '')
    if not username or not password or role not in ('業務', '會計'):
        flash("請填寫完整資料", "danger")
        return redirect(url_for('settings'))
    if len(password) < 4:
        flash("密碼至少需要 4 個字元", "danger")
        return redirect(url_for('settings'))
    try:
        db.create_user(request.form.to_dict())
        flash(f"帳號「{username}」已新增！", "success")
    except Exception:
        flash("帳號名稱已存在，請換一個", "danger")
    return redirect(url_for('settings'))


@app.route("/settings/users/<int:user_id>/edit", methods=["POST"])
def settings_user_edit(user_id):
    if session.get('role') != '老闆':
        return redirect(url_for('settings'))
    db.update_user(user_id, request.form.to_dict())
    flash("帳號已更新！", "success")
    return redirect(url_for('settings'))


@app.route("/settings/users/<int:user_id>/delete", methods=["POST"])
def settings_user_delete(user_id):
    if session.get('role') != '老闆':
        return redirect(url_for('settings'))
    if user_id == session['user_id']:
        flash("不能刪除自己的帳號", "danger")
        return redirect(url_for('settings'))
    db.delete_user(user_id)
    flash("帳號已刪除。", "success")
    return redirect(url_for('settings'))


# ── 黑名單 ────────────────────────────────────────────

@app.route("/blacklist")
def blacklist():
    search = request.args.get("q", "")
    rows = db.list_blacklist(search or None)
    return render_template("blacklist.html", rows=rows, search=search)


@app.route("/blacklist/new", methods=["POST"])
def blacklist_new():
    if session.get('role') not in ('老闆', '會計'):
        return redirect(url_for('blacklist'))
    db.create_blacklist(request.form.to_dict())
    flash("已加入黑名單！", "success")
    return redirect(url_for('blacklist'))


@app.route("/blacklist/<int:bid>/delete", methods=["POST"])
def blacklist_delete(bid):
    if session.get('role') not in ('老闆', '會計'):
        return redirect(url_for('blacklist'))
    db.delete_blacklist(bid)
    flash("已從黑名單移除。", "success")
    return redirect(url_for('blacklist'))


@app.route("/blacklist/import", methods=["POST"])
def blacklist_import():
    if session.get('role') not in ('老闆', '會計'):
        return redirect(url_for('blacklist'))
    import io
    import openpyxl
    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        flash("請上傳 .xlsx 格式的 Excel 檔案", "danger")
        return redirect(url_for('blacklist'))
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        ok = skip = 0
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:
                continue  # 跳過標題列
            name = str(row[0]).strip() if row[0] else ""
            if not name or name == "None":
                skip += 1
                continue
            id_card = str(row[1]).strip() if row[1] else ""
            phone   = str(row[2]).strip() if row[2] else ""
            reason  = str(row[3]).strip() if row[3] else ""
            notes   = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            if reason not in ('詢問', '呆帳'):
                reason = '詢問'
            db.create_blacklist({'name': name, 'id_card': id_card,
                                 'phone': phone, 'reason': reason, 'notes': notes})
            ok += 1
        flash(f"匯入完成：成功 {ok} 筆，略過空白列 {skip} 筆。", "success")
    except Exception as e:
        flash(f"匯入失敗：{e}", "danger")
    return redirect(url_for('blacklist'))


# ── 啟動 ──────────────────────────────────────────────

def open_browser(port):
    webbrowser.open(f"http://127.0.0.1:{port}")


if __name__ == "__main__":
    db.init_db()
    port = 5678
    # debug=True 讓程式碼和模板變更後自動重載，F5 即可看到效果
    # WERKZEUG_RUN_MAIN 判斷避免瀏覽器開兩次
    is_frozen = getattr(sys, 'frozen', False)
    if not os.environ.get("NO_BROWSER") and not os.environ.get("WERKZEUG_RUN_MAIN"):
        threading.Timer(1.2, open_browser, args=[port]).start()
    app.run(host="127.0.0.1", port=port, debug=not is_frozen, use_reloader=not is_frozen)
